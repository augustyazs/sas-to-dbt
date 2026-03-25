from pathlib import Path
from models.schemas import DbtProject


def write_dbt_project(project: DbtProject, output_dir: Path) -> list[str]:
    """Write all dbt project files to disk."""
    output_dir.mkdir(parents=True, exist_ok=True)
    written = []

    file_map = {
        "dbt_project.yml": project.dbt_project_yml,
        "models/sources.yml": project.sources_yml,
        "models/schema.yml": project.schema_yml,
    }

    for rel_path, content in file_map.items():
        if content and content.strip():
            fp = output_dir / rel_path
            fp.parent.mkdir(parents=True, exist_ok=True)
            fp.write_text(content, encoding="utf-8")
            written.append(str(fp))

    for model in project.models:
        fp = output_dir / model.path
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(model.content, encoding="utf-8")
        written.append(str(fp))

    for macro in project.macros:
        fp = output_dir / macro.path
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(macro.content, encoding="utf-8")
        written.append(str(fp))

    if project.not_converted:
        fp = output_dir / "NOT_CONVERTED.md"
        lines = ["# Blocks Not Converted to dbt\n"]
        for item in project.not_converted:
            lines.append(f"- {item}\n")
        fp.write_text("\n".join(lines), encoding="utf-8")
        written.append(str(fp))

    return written


def write_sas_documentation(documentation: str, output_dir: Path) -> str:
    """Write SAS source documentation as a Markdown file."""
    output_dir.mkdir(parents=True, exist_ok=True)
    fp = output_dir / "sas_documentation.md"
    fp.write_text(documentation, encoding="utf-8")
    return str(fp)


def write_sttm_excel(sttm_data: dict, output_dir: Path) -> str:
    """Write STTM data to a formatted Excel file with one tab per output table."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output_dir.mkdir(parents=True, exist_ok=True)
    fp = output_dir / "sttm.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADERS = [
        "Target Schema", "Target Table", "Target Column", "Target Data Type",
        "Transformation Rule",
        "Source Schema", "Source Table", "Source Column", "Source Data Type",
        "Additional Comments",
    ]
    header_font      = Font(bold=True, color="FFFFFF", size=11)
    header_fill      = PatternFill("solid", fgColor="1F4E79")
    alt_fill         = PatternFill("solid", fgColor="DCE6F1")
    wrap_alignment   = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="top")
    thin_border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    col_widths = [18, 22, 25, 18, 45, 18, 22, 25, 18, 40]

    for tab in sttm_data.get("tabs", []):
        tab_name = tab.get("tab_name", "Output")[:31]
        ws = wb.create_sheet(title=tab_name)

        desc = tab.get("description", "")
        if desc:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
            desc_cell = ws.cell(row=1, column=1, value=desc)
            desc_cell.font      = Font(italic=True, size=10, color="444444")
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[1].height = 30
            header_row = 2
        else:
            header_row = 1

        for ci, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=header_row, column=ci, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center_alignment; c.border = thin_border
        ws.row_dimensions[header_row].height = 20

        for ri, row in enumerate(tab.get("rows", []), start=header_row + 1):
            vals = [
                row.get("target_schema", ""), row.get("target_table", ""),
                row.get("target_column", ""), row.get("target_data_type", ""),
                row.get("transformation_rule", ""), row.get("source_schema", ""),
                row.get("source_table", ""), row.get("source_column", ""),
                row.get("source_data_type", ""), row.get("additional_comments", ""),
            ]
            fill = alt_fill if (ri - header_row) % 2 == 0 else None
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.alignment = wrap_alignment; c.border = thin_border
                if fill: c.fill = fill
            ws.row_dimensions[ri].height = 40

        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(fp)
    return str(fp)